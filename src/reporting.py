# src/reporting.py
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone

import discord
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from repo.report_repo import list_items_for_report, list_movements_in_epoch_range, delete_movements_before_epoch
from repo.settings_repo import get_settings, update_settings
from utils.time_kst import now_kst


KST = timezone(timedelta(hours=9))


def _style_header(ws, header_row=1):
    bold = Font(bold=True)
    for cell in ws[header_row]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = f"A{header_row + 1}"


def _autosize(ws, max_col: int):
    # 대충 보기 좋은 폭
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def _wb_to_file(wb: Workbook, filename: str) -> discord.File:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return discord.File(fp=bio, filename=filename)


def _action_kor(action: str) -> str:
    return {"IN": "입고", "OUT": "출고", "ADJUST": "정정"}.get(action, action)


def build_daily_inventory_wb(conn, guild_id: int) -> Workbook:
    items = list_items_for_report(conn, guild_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "일일 재고 보고서"

    ws.append(["카테고리", "품목명", "코드", "현재재고", "경고기준", "보관 위치", "메모", "상태"])
    for it in items:
        ws.append([
            it.get("category_name") or "기타",
            it.get("name") or "",
            it.get("code") or "",
            it.get("qty"),
            it.get("warn_below"),
            it.get("storage_location") or "",
            it.get("note") or "",
            "활성" if int(it.get("is_active", 1)) == 1 else "비활성",
        ])

    _style_header(ws)
    _autosize(ws, 8)
    return wb


def build_daily_log_wb(conn, guild_id: int, start_epoch: int, end_epoch: int) -> Workbook:
    rows = list_movements_in_epoch_range(conn, guild_id, start_epoch, end_epoch)

    wb = Workbook()
    ws = wb.active
    ws.title = "일일 로그 기록"

    # ✅ 요약 계산
    total_in = 0
    total_out = 0
    adj_plus = 0
    adj_minus = 0
    for r in rows:
        act = str(r.get("action") or "")
        q = int(r.get("qty_change") or 0)
        if act == "IN":
            total_in += q
        elif act == "OUT":
            total_out += abs(q)
        elif act == "ADJUST":
            if q >= 0:
                adj_plus += q
            else:
                adj_minus += abs(q)

    # ✅ 요약 1줄 (A1~J1 병합)
    summary = f"요약: 총 입고 {total_in} · 총 출고 {total_out} · 정정 +{adj_plus}/-{adj_minus} · 로그 {len(rows)}건"
    ws.append([summary])
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(vertical="center")

    # ✅ 헤더는 2행
    ws.append(["시간(KST)", "작업", "카테고리", "품목명", "코드", "변동수량", "재고(전)", "재고(후)", "사유", "수정자"])

    for r in rows:
        act = str(r.get("action") or "")
        qty_change = int(r.get("qty_change") or 0)

        # ✅ 변동수량 표시(출고도 양수, 정정만 부호)
        if act == "ADJUST":
            sign = "+" if qty_change >= 0 else ""
            change_text = f"{sign}{qty_change}"
        else:
            change_text = str(abs(qty_change))

        ws.append([
            r.get("created_at_kst_text"),
            _action_kor(act),
            r.get("category_name_snapshot") or "",
            r.get("item_name_snapshot") or "",
            r.get("item_code_snapshot") or "",
            change_text,
            r.get("before_qty"),
            r.get("after_qty"),
            r.get("reason") or "",
            r.get("discord_name") or "",
        ])

    _style_header(ws, header_row=2)
    _autosize(ws, 10)
    return wb


def build_monthly_log_wb(conn, guild_id: int, start_epoch: int, end_epoch: int, ym_text: str) -> Workbook:
    rows = list_movements_in_epoch_range(conn, guild_id, start_epoch, end_epoch)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "월간 누적 로그"

    # ✅ 요약 계산
    total_in = 0
    total_out = 0
    adj_plus = 0
    adj_minus = 0
    for r in rows:
        act = str(r.get("action") or "")
        q = int(r.get("qty_change") or 0)
        if act == "IN":
            total_in += q
        elif act == "OUT":
            total_out += abs(q)
        elif act == "ADJUST":
            if q >= 0:
                adj_plus += q
            else:
                adj_minus += abs(q)

    summary = f"요약: 총 입고 {total_in} · 총 출고 {total_out} · 정정 +{adj_plus}/-{adj_minus} · 로그 {len(rows)}건"
    ws1.append([summary])
    ws1.merge_cells("A1:J1")
    ws1["A1"].font = Font(bold=True)
    ws1["A1"].alignment = Alignment(vertical="center")

    # ✅ 헤더는 2행
    ws1.append(["시간(KST)", "작업", "카테고리", "품목명", "코드", "변동수량", "재고(전)", "재고(후)", "사유", "수정자"])

    for r in rows:
        act = str(r.get("action") or "")
        qty_change = int(r.get("qty_change") or 0)

        if act == "ADJUST":
            sign = "+" if qty_change >= 0 else ""
            change_text = f"{sign}{qty_change}"
        else:
            change_text = str(abs(qty_change))

        ws1.append([
            r.get("created_at_kst_text"),
            _action_kor(act),
            r.get("category_name_snapshot") or "",
            r.get("item_name_snapshot") or "",
            r.get("item_code_snapshot") or "",
            change_text,
            r.get("before_qty"),
            r.get("after_qty"),
            r.get("reason") or "",
            r.get("discord_name") or "",
        ])

    _style_header(ws1, header_row=2)
    _autosize(ws1, 10)

    # 간단 요약 시트(품목별 IN/OUT 합)
    ws2 = wb.create_sheet("요약")
    ws2.append(["품목명", "코드", "총 입고", "총 출고", "정정 합계"])
    summary = {}
    for r in rows:
        key = (r.get("item_name_snapshot") or "", r.get("item_code_snapshot") or "")
        s = summary.setdefault(key, {"IN": 0, "OUT": 0, "ADJUST": 0})
        act = str(r.get("action") or "")
        s[act] = s.get(act, 0) + int(r.get("qty_change") or 0)

    for (name, code), s in summary.items():
        ws2.append([name, code, s.get("IN", 0), abs(s.get("OUT", 0)), s.get("ADJUST", 0)])
    _style_header(ws2)
    _autosize(ws2, 5)

    return wb


async def _get_report_channel(interaction_client, guild: discord.Guild):
    conn = interaction_client.conn
    s = get_settings(conn, guild.id)
    ch_id = s.get("report_channel_id") or s.get("alert_channel_id")
    if not ch_id:
        return None
    ch = guild.get_channel(int(ch_id))
    return ch if isinstance(ch, discord.TextChannel) else None


def _kst_day_range_epochs(dt_kst: datetime) -> tuple[int, int]:
    start = dt_kst.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start + timedelta(days=1)
    return int(start.timestamp()), int(end.timestamp())


def _kst_month_range_epochs(dt_kst: datetime) -> tuple[int, int]:
    start = dt_kst.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    # 다음달 1일
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)
    return int(start.timestamp()), int(end.timestamp())


def _quarter_key(dt_kst: datetime) -> str:
    q = ((dt_kst.month - 1) // 3) + 1
    return f"{dt_kst.year}-Q{q}"


def _start_of_current_quarter(dt_kst: datetime) -> datetime:
    q = ((dt_kst.month - 1) // 3) + 1
    start_month = 1 + (q - 1) * 3
    return dt_kst.replace(month=start_month, day=1, hour=0, minute=0, second=0, microsecond=0)


async def run_daily_reports(client, guild: discord.Guild):
    conn = client.conn
    s = get_settings(conn, guild.id)

    # 보고서 시간
    h = int(s.get("report_hour", 18))
    m = int(s.get("report_minute", 30))

    k = now_kst()
    dt = k.dt  # KST aware datetime (초=0 권장)

    today = dt.strftime("%Y-%m-%d")
    last_done = (s.get("last_daily_report_date") or "")

    # ✅ 오늘의 스케줄 시각(18:30)
    scheduled = dt.replace(hour=h, minute=m, second=0, microsecond=0)

    # ✅ 핵심: 18:30 "이후"이고 오늘 아직 안 올렸으면 올리기
    if dt < scheduled:
        return
    if last_done == today:
        return

    ch = await _get_report_channel(client, guild)
    if not ch:
        return

    # 오늘 00:00~24:00 범위
    start_epoch, end_epoch = _kst_day_range_epochs(dt)

    wb_inv = build_daily_inventory_wb(conn, guild.id)
    wb_log = build_daily_log_wb(conn, guild.id, start_epoch, end_epoch)

    date_text = dt.strftime("%Y-%m-%d")
    f1 = _wb_to_file(wb_inv, f"일일_재고보고서_{date_text}.xlsx")
    f2 = _wb_to_file(wb_log, f"일일_로그기록_{date_text}.xlsx")



    await ch.send(content=f"📌 일일 보고서 / 로그 ({dt.strftime('%Y/%m/%d')})", files=[f1, f2])

    update_settings(conn, guild.id, last_daily_report_date=today)

    # ✅ 월간 누적 업로드(말일 놓침 대비)
    # - "오늘이 1일이고 18:30 이후"면 지난 달 월간을 올린다.
    if dt.day == 1:
        # 지난 달 YYYY-MM
        prev_month = (dt.replace(day=1) - timedelta(days=1))
        ym = prev_month.strftime("%Y-%m")
        if (s.get("last_monthly_report_ym") or "") != ym:
            ms, me = _kst_month_range_epochs(prev_month)
            wb_month = build_monthly_log_wb(conn, guild.id, ms, me, ym)
            month_text = prev_month_dt.strftime("%Y-%m")
            fm = _wb_to_file(wb_month, f"월간_누적로그_{month_text}.xlsx")
            await ch.send(content=f"📚 월간 누적 로그 ({ym})", file=fm)
            update_settings(conn, guild.id, last_monthly_report_ym=ym)

    # - 추가로: 말일 당일에 살아있으면 그날도 올리고 싶다? -> 원하면 여기서 “말일이면 바로”도 가능


async def run_quarterly_cleanup(client, guild: discord.Guild):
    """
    분기마다(3개월) '옛날 기록(= movements)'만 삭제.
    기준: 현재 분기 시작일 00:00 이전 데이터는 삭제.

    ✅ 놓침 대비:
    - 분기 첫날 00:05를 놓쳐도
    - 분기 첫 주(day 1~7) 중 아무 때나 1회 실행
    """
    conn = client.conn
    s = get_settings(conn, guild.id)

    k = now_kst()
    dt = k.dt

    # 분기 첫 주에만 시도 (너무 넓히고 싶으면 10일로 늘려도 됨)
    if dt.day > 7:
        return

    qkey = _quarter_key(dt)
    if (s.get("last_quarter_cleanup") or "") == qkey:
        return

    # 분기 시작일(이번 분기)
    cutoff_dt = _start_of_current_quarter(dt)
    cutoff_epoch = int(cutoff_dt.timestamp())

    deleted = delete_movements_before_epoch(conn, guild.id, cutoff_epoch)

    ch = await _get_report_channel(client, guild)
    if ch:
        await ch.send(
            f"🧹 분기 로그 정리 완료: {deleted}건 삭제 "
            f"(기준: {cutoff_dt.strftime('%Y/%m/%d %H:%M:%S')} KST 이전)"
        )

    update_settings(conn, guild.id, last_quarter_cleanup=qkey)

async def force_send_daily_reports(client, guild: discord.Guild, mark_done: bool = True) -> bool:
    """
    ✅ 지금 즉시 '오늘자' 일일 재고보고서 + 일일 로그 업로드
    - mark_done=True면 오늘 스케줄 업로드도 중복되지 않게 last_daily_report_date 기록
    """
    conn = client.conn
    ch = await _get_report_channel(client, guild)
    if not ch:
        return False

    k = now_kst()
    dt = k.dt  # 오늘(KST)

    start_epoch, end_epoch = _kst_day_range_epochs(dt)

    wb_inv = build_daily_inventory_wb(conn, guild.id)
    wb_log = build_daily_log_wb(conn, guild.id, start_epoch, end_epoch)

    date_text = dt.strftime("%Y-%m-%d")
    f1 = _wb_to_file(wb_inv, f"일일_재고보고서_{date_text}.xlsx")
    f2 = _wb_to_file(wb_log, f"일일_로그기록_{date_text}.xlsx")


    await ch.send(content=f"📌 (수동) 일일 보고서 / 로그 ({dt.strftime('%Y/%m/%d')})", files=[f1, f2])

    if mark_done:
        today = dt.strftime("%Y-%m-%d")
        update_settings(conn, guild.id, last_daily_report_date=today)

    return True


async def force_send_monthly_prev_month(client, guild: discord.Guild, mark_done: bool = True) -> bool:
    """
    ✅ 지금 즉시 '지난달' 월간 누적 로그 업로드
    - mark_done=True면 동일 월 중복 업로드 방지(last_monthly_report_ym 기록)
    """
    conn = client.conn
    ch = await _get_report_channel(client, guild)
    if not ch:
        return False

    k = now_kst()
    dt = k.dt

    prev_month_dt = (dt.replace(day=1) - timedelta(days=1))
    ym = prev_month_dt.strftime("%Y-%m")

    ms, me = _kst_month_range_epochs(prev_month_dt)
    wb_month = build_monthly_log_wb(conn, guild.id, ms, me, ym)
    month_text = prev_month_dt.strftime("%Y-%m")
    fm = _wb_to_file(wb_month, f"월간_누적로그_{month_text}.xlsx")


    await ch.send(content=f"📚 (수동) 월간 누적 로그 ({ym})", file=fm)

    if mark_done:
        update_settings(conn, guild.id, last_monthly_report_ym=ym)

    return True